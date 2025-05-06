import pandas as pd
import openpyxl
from PyQt5 import uic
from PyQt5.QtCore import QAbstractTableModel, Qt
from PyQt5.QtWidgets import QWidget, QTableView, QHeaderView, QFileDialog

from openpyxl.styles import Font, PatternFill
from controllers.transport_controller import TransportController
from controllers.fresh_air_controller import FreshAirController
from controllers.vibration_controller import VibrationController
from controllers.hopper_heater_controller import HopperHeaterController
from views.message_box_view import show_message


class ResultTab(QWidget):
    def __init__(self, main_view, project_details):
        super().__init__()
        uic.loadUi("ui/results_tab.ui", self)

        self.main_view = main_view
        self.project_details = project_details

        self.tables = {
            "transport_panel_table": self.transport_panel_table,
            "fresh_air_panel_table": self.fresh_air_panel_table,
            "vibration_panel_table": self.vibration_panel_table,
            "hopper_heater_panel_table": self.hopper_heater_panel_table,
        }

        self.panels = {}

        self._setup_result_table()

        self.excel_btn.clicked.connect(self._export_to_excel)
        self.show()

    def _setup_result_table(self):
        for table in self.tables.values():
            table.setAlternatingRowColors(True)
            table.setHorizontalScrollMode(QTableView.ScrollPerPixel)
            table.setVerticalScrollMode(QTableView.ScrollPerPixel)
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setWordWrap(True)
            table.setTextElideMode(Qt.ElideRight)
            table.verticalHeader().setVisible(False)

    def generate_panels(self):
        transport_controller = TransportController()
        self.panels["transport_panel"] = transport_controller.build_panel(self.project_details)
        fresh_air_controller = FreshAirController()
        self.panels["fresh_air_panel"] = fresh_air_controller.build_panel(self.project_details)
        vibration_controller = VibrationController()
        self.panels["vibration_panel"] = vibration_controller.build_panel(self.project_details)
        hopper_heater_controller = HopperHeaterController()
        self.panels["hopper_heater_panel"] = hopper_heater_controller.build_panel(self.project_details)

        self.generate_table(self.panels["transport_panel"], self.tables["transport_panel_table"])
        self.generate_table(self.panels["fresh_air_panel"], self.tables["fresh_air_panel_table"])
        self.generate_table(self.panels["vibration_panel"], self.tables["vibration_panel_table"])
        self.generate_table(self.panels["hopper_heater_panel"], self.tables["hopper_heater_panel_table"])

    def generate_table(self, panel, table):
        df = pd.DataFrame(panel)
        df = self._add_summary_row(df)
        model = PandasModel(df)
        table.setModel(model)

        self._resize_columns_to_contents(model, table)

        header = table.horizontalHeader()
        for col in range(model.columnCount(None) - 1):
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(model.columnCount(None) - 1, QHeaderView.Stretch)

        table.resizeRowsToContents()

    def _resize_columns_to_contents(self, model, table):
        metrics = table.fontMetrics()

        for col in range(model.columnCount(None)):
            max_width = metrics.horizontalAdvance(str(model.headerData(col, Qt.Horizontal, Qt.DisplayRole))) + 20
            for row in range(model.rowCount(None)):
                index = model.index(row, col)
                text = str(model.data(index, Qt.DisplayRole))
                max_width = max(max_width, metrics.horizontalAdvance(text) + 20)
            table.setColumnWidth(col, max_width)

    def _add_summary_row(self, df):
        summary = {
            col: df[col].sum() if col == "total_price" else ("Total" if col.lower() == "type" else "")
            for col in df.columns
        }
        return pd.concat([df, pd.DataFrame([summary], index=["Total"])])

    def _export_to_excel(self):

        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "ReportByGriinPower",
                                                   "Excel Files (*.xlsx)")
        if not file_path:
            return

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for name, table in self.tables.items():
                model = table.model()
                if model is None:
                    continue

                df = model._data.copy()
                df.index = range(1, len(df) + 1)

                df.to_excel(writer, sheet_name=name, index=True, startrow=1)

                workbook = writer.book
                worksheet = writer.sheets[name]

                # Styles
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                center_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
                zebra_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
                total_fill = PatternFill(start_color="FFFFAA00", end_color="FFFFAA00", fill_type="solid")

                columns = list(df.columns.insert(0, df.index.name or ""))  # Including index

                # Format headers
                for col_idx, col_name in enumerate(columns, 1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # Format data rows
                for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    row_idx = row[0].row
                    for col_idx, cell in enumerate(row, 0):
                        cell.alignment = center_alignment
                        cell.border = thin_border

                        col_name = columns[col_idx]
                        if col_name.lower() in ("price", "total_price") and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

                    if row_idx % 2 == 1:
                        for cell in row:
                            cell.fill = zebra_fill

                # Special styling for Total row
                total_row = worksheet.max_row
                for cell in worksheet[total_row]:
                    cell.fill = total_fill

                # Auto-fit column widths
                for column_cells in worksheet.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    worksheet.column_dimensions[column].width = max(max_length + 2, 10)

                # Freeze header row
                worksheet.freeze_panes = worksheet["A3"]


class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data.copy()

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or role not in (Qt.DisplayRole, Qt.EditRole):
            return None

        value = self._data.iat[index.row(), index.column()]
        column_name = self._data.columns[index.column()]

        if column_name in {"price", "total_price"} and pd.notnull(value):
            try:
                value = float(value)
                return f"{value:,.0f}" if value.is_integer() else f"{value:,.2f}"
            except (ValueError, TypeError):
                return str(value)

        return str(value)

    def setData(self, index, value, role=Qt.EditRole):
        if not index.isValid() or role != Qt.EditRole:
            return False

        try:
            row, col = index.row(), index.column()
            dtype = self._data.dtypes[col]
            self._data.iat[row, col] = dtype.type(value)
            self.dataChanged.emit(index, index)
            return True
        except Exception as e:
            show_message(f"Failed to set data at ({row}, {col}): {e}", "Error")
            return False

    def flags(self, index):
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self._data.columns[section].replace("_", " ").upper()
        if orientation == Qt.Vertical:
            return str(self._data.index[section])
        return None
