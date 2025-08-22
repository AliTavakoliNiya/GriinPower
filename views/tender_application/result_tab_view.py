import copy
import subprocess
from typing import Optional, Tuple

import pandas as pd
from PyQt5 import uic
from PyQt5.QtCore import QAbstractTableModel, Qt, QModelIndex
from PyQt5.QtGui import QPainter
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtWidgets import (
    QComboBox,
    QFileDialog,
    QHeaderView,
    QMainWindow,
    QStyledItemDelegate,
    QTableView,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget, QPushButton,
)

from controllers.tender_application.bagfilter_controller import BagfilterController
from controllers.tender_application.cable_controller import CableController
from controllers.tender_application.electric_motor_controller import ElectricMotorController
from controllers.tender_application.fan_damper_controller import FanDamperController
from controllers.tender_application.fresh_air_controller import FreshAirController
from controllers.tender_application.hopper_heater_controller import HopperHeaterController
from controllers.tender_application.installation_controller import InstallationController
from controllers.tender_application.project_session_controller import ProjectSession
from controllers.tender_application.transport_controller import TransportController
from controllers.tender_application.vibration_controller import VibrationController
from utils.pandas_model import PandasModel
from views.message_box_view import show_message


class ResultTab(QWidget):
    """
    ResultTab: Renders panel data into tables, generates a summary table,
    and supports Excel export and a details tree viewer.

    NOTE:
    - Paths and UI object names are preserved.
    - Logic flow is unchanged; only safe-guards, readability, and comments added.
    """

    def __init__(self, main_view: QWidget):
        super().__init__()
        # Keep original, project-relative UI path as requested
        uic.loadUi("ui/tender_application/results_tab.ui", self)

        self.main_view = main_view
        self.current_project = ProjectSession()
        self.electrical_specs = self.current_project.project_electrical_specs

        self._init_tables()
        self._init_tabs()
        self._style_all_tables()

        # Wire up UI actions
        self.excel_btn.clicked.connect(self._export_to_excel)
        self.show_datail_btn.clicked.connect(self._show_detail_tree)
        self.update_table.clicked.connect(self.generate_data)

    # -------------------------- Setup helpers --------------------------

    def _init_tables(self) -> None:
        """Map all known table objectNames to their widgets (UI names unchanged)."""
        self.tables = {
            "bagfilter_table": self.bagfilter_table,
            "fan_damper_table": self.fan_damper_table,
            "transport_table": self.transport_table,
            "fresh_air_table": self.fresh_air_table,
            "vibration_table": self.vibration_table,
            "hopper_heater_table": self.hopper_heater_table,
            "cable_table": self.cable_table,
            "installation_table": self.installation_table,
            "summary_table": self.summary_table,
        }

    def _init_tabs(self) -> None:
        """Keep a list of (table_name, tab_widget, label) to rebuild visible tabs."""
        self.all_tabs = [
            ("bagfilter_table", self.bagfilter_tab, "Bagfilter"),
            ("fan_damper_table", self.fan_damper_tab, "Fan Damper"),
            ("transport_table", self.transport_tab, "Transport"),
            ("fresh_air_table", self.fresh_air_tab, "Fresh Air"),
            ("vibration_table", self.vibration_tab, "Vibration"),
            ("hopper_heater_table", self.hopper_heater_tab, "Hopper Heater"),
            ("cable_table", self.cable_tab, "Cable"),
            ("installation_table", self.installation_tab, "Installation"),
            ("summary_table", self.summary_tab, "Summary"),
        ]

    def _style_all_tables(self) -> None:
        """Apply consistent table styling and behavior."""
        for table in self.tables.values():
            table.setAlternatingRowColors(True)
            table.setHorizontalScrollMode(QTableView.ScrollPerPixel)
            table.setVerticalScrollMode(QTableView.ScrollPerPixel)
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setWordWrap(True)
            table.setTextElideMode(Qt.ElideRight)
            table.verticalHeader().setVisible(False)
            table.horizontalHeader().setVisible(True)

    # -------------------------- UI actions --------------------------

    def _show_detail_tree(self) -> None:
        """Open a simple dictionary viewer for the project's electrical specs."""
        viewer = DictionaryTreeViewer(data=self.electrical_specs, parent=self.main_view)
        viewer.show()

    def generate_data(self) -> None:
        """
        Build all panels via controllers, populate the corresponding tables,
        and then build the summary table (with selectable motor brand if fan is enabled).
        """
        self.panels = self._build_all_panels()

        # Populate each individual panel table
        for key, panel in self.panels.items():
            table_name = key.replace("_panel", "_table")
            if table_name in self.tables:
                self._populate_table_view(panel, self.tables[table_name])

        panels_copy = copy.deepcopy(self.panels)

        # Summary generation
        if self.main_view.electrical_tab.fan_checkbox.isChecked():
            motor_ctrl = ElectricMotorController()
            motor_info = motor_ctrl.calculate_price()
            summary_data = self._generate_summary_data(panels_copy, motor_info)
            # Use a model that shows a drop-down of motor brands in the "Note" column.
            self._set_summary_model(summary_data, use_brands=True)
        else:
            summary_data = self._generate_summary_data(panels_copy)
            self._populate_table_view(summary_data, self.tables["summary_table"])

        # Show only tabs that contain non-zero totals
        self._refresh_tabs()

    # -------------------------- Data builders --------------------------

    def _build_all_panels(self) -> dict:
        """Call each controller's build_panel and return a mapping of panel_key -> list[dict]."""
        controllers = {
            "bagfilter_panel": BagfilterController,
            "fan_damper_panel": FanDamperController,
            "transport_panel": TransportController,
            "fresh_air_panel": FreshAirController,
            "vibration_panel": VibrationController,
            "hopper_heater_panel": HopperHeaterController,
            "cable_panel": CableController,
            "installation_panel": InstallationController,
        }
        return {key: ctrl().build_panel() for key, ctrl in controllers.items()}

    def _populate_table_view(self, panel, table: QTableView) -> None:
        """
        Convert panel (list of rows or dict-of-lists) to DataFrame,
        append a summary row, put into a PandasModel, and set on the table.
        """
        df = pd.DataFrame(panel)
        df = self._add_summary_row(df)
        model = PandasModel(df)
        table.setModel(model)
        self._resize_columns(table, model)

    def _resize_columns(self, table: QTableView, model: QAbstractTableModel) -> None:
        """
        Let Qt compute sensible widths for all columns; last column stretches.
        This avoids an expensive per-cell measurement loop.
        """
        header: QHeaderView = table.horizontalHeader()
        col_count = model.columnCount(None)
        if col_count <= 0:
            return
        for col in range(col_count - 1):
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(col_count - 1, QHeaderView.Stretch)
        table.resizeRowsToContents()

    def _add_summary_row(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Append a summary row:
        - If 'total_price' exists, sum it (numeric coercion).
        - If a 'type' column exists (case-insensitive), place the label 'Total' there.
        Always returns a DataFrame (original + one extra row) unless df is empty.
        """
        if df is None or df.empty:
            return df if df is not None else pd.DataFrame()

        # Normalize columns for case-insensitive lookup
        lower_cols = {c.lower(): c for c in df.columns}

        # Prepare summary template
        summary = {col: "" for col in df.columns}

        # Sum total_price if present
        if "total_price" in lower_cols:
            price_col = lower_cols["total_price"]
            summary[price_col] = (
                pd.to_numeric(df[price_col], errors="coerce").fillna(0).sum()
            )

        # Put the label 'Total' in 'type' if present
        if "type" in lower_cols:
            type_col = lower_cols["type"]
            summary[type_col] = "Total"

        return pd.concat([df, pd.DataFrame([summary], index=["Total"])])

    def _refresh_tabs(self) -> None:
        """
        Rebuild the tab bar based on whether each table has a non-zero total.
        Supports:
          • PandasModel with `total_price` (standard panels)
          • PandasModel with `Price` for the summary_table (fan unchecked path)
          • SummaryTableModel (list[dict]) with a 'Price' total in the last row
        """
        self.tabWidget.clear()

        for table_name, tab_widget, label in self.all_tabs:
            table = self.tables.get(table_name)
            if table is None:
                continue

            model = table.model()
            if model is None:
                continue

            df_or_list = getattr(model, "_data", None)

            # ---- PandasModel case: DataFrame ----
            if isinstance(df_or_list, pd.DataFrame):
                try:
                    df = df_or_list
                    if df.empty:
                        continue

                    # Normal rule for panels: use 'total_price' in the last row
                    if "total_price" in df.columns:
                        last_val = pd.to_numeric(df.iloc[-1]["total_price"], errors="coerce")
                        if pd.notna(last_val) and float(last_val) > 0:
                            self.tabWidget.addTab(tab_widget, label)
                        continue

                    # SPECIAL CASE: summary_table when fan is unchecked -> uses 'Price'
                    if table_name == "summary_table" and "Price" in df.columns:
                        last_val = pd.to_numeric(df.iloc[-1]["Price"], errors="coerce")
                        if pd.notna(last_val) and float(last_val) > 0:
                            self.tabWidget.addTab(tab_widget, label)
                except Exception:
                    # Be defensive: if parsing fails, skip showing this tab.
                    pass
                continue

            # ---- SummaryTableModel case: list[dict] ----
            if isinstance(df_or_list, list) and df_or_list:
                try:
                    last_row = df_or_list[-1]
                    total_price = float(last_row.get("Price", 0))
                    if total_price > 0:
                        self.tabWidget.addTab(tab_widget, label)
                except Exception:
                    pass

    def _should_show_tab(self, df) -> bool:
        """
        Legacy helper retained for compatibility. Prefer _refresh_tabs logic above.
        Kept behavior unchanged but guarded to avoid KeyError on empty data.
        """
        if isinstance(df, pd.DataFrame):
            if df.empty or "total_price" not in df.columns:
                return False
            try:
                return float(df.iloc[-1]["total_price"]) > 0
            except Exception:
                return False
        elif isinstance(df, list) and df and isinstance(df[-1], dict):
            try:
                return float(df[-1].get("Price", 0)) > 0
            except Exception:
                return False
        return False

    def _generate_summary_data(self, panels: dict, electric_motor_price_and_info=None) -> dict:
        """
        Build summary rows from panel totals, optionally with electric motor choices.
        Returns dict-of-lists compatible with DataFrame construction.
        """
        summary = {"Title": [], "Price": [], "Note": [], "brands": []}
        total = 0.0

        # Accumulate per-panel totals
        for name, panel in panels.items():
            df = pd.DataFrame(panel)
            panel_total = 0.0
            if not df.empty and "total_price" in df.columns:
                panel_total = pd.to_numeric(df["total_price"], errors="coerce").fillna(0).sum()
            if panel_total == 0:
                continue
            summary["Title"].append(name.replace("_", " ").title())
            summary["Price"].append(float(panel_total))
            summary["Note"].append("")
            summary["brands"].append({})
            total += float(panel_total)

        # Optionally include Electric Motor row (if fan is enabled and info is present)
        fan_enabled = bool(self.electrical_specs.get("fan", {}).get("status"))
        if fan_enabled and electric_motor_price_and_info:
            # Merge brand options from motor info into a single dict
            merged_brands = {}
            for motor in electric_motor_price_and_info:
                merged_brands.update(motor.get("brands", {}))

            # Choose a default brand (first item) if available
            default_note, default_price = next(iter(merged_brands.items()), ("", 0))

            summary["Title"].append("Electric Motor")
            summary["Price"].append(float(default_price))
            summary["Note"].append(default_note)
            summary["brands"].append(merged_brands)
            total += float(default_price)

        # Grand Total
        summary["Title"].append("Total")
        summary["Price"].append(float(total))
        summary["Note"].append("")
        summary["brands"].append({})


        return summary

    def _set_summary_model(self, summary_data: dict, use_brands: bool = False) -> None:
        """
        Set an interactive summary model on the summary table.
        When use_brands=True, a combo delegate allows brand selection in the Note column.
        """
        rows = pd.DataFrame(summary_data).to_dict(orient="records")
        model = SummaryTableModel(rows)
        table = self.tables["summary_table"]
        table.setModel(model)
        self._resize_columns(table, model)
        if use_brands:
            table.setItemDelegateForColumn(2, BrandComboDelegate(table))

    # -------------------------- Export to Excel --------------------------

    def _export_to_excel(self) -> None:
        """
        Handle user-initiated export to Excel with full report and factor sheets.
        Keeps original naming scheme for the save dialog default.
        """
        file_name = (
            f"{self.current_project.name}-"
            f"{self.current_project.code}-"
            f"{self.current_project.unique_no}-"
            f"Rev{str(self.current_project.revision).zfill(2)}"
        )
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", file_name, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            # NOTE: The code previously computed 'report_path' but never wrote to it.
            # To avoid opening a non-existent file, we open the actual 'file_path' after writing.
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Export main report sheets
                for name, table in self.tables.items():
                    model = table.model()
                    if model is None:
                        continue

                    df = getattr(model, "_data", None)
                    if df is None:
                        continue
                    if isinstance(df, list):
                        df = pd.DataFrame(df)
                    elif not isinstance(df, pd.DataFrame):
                        continue

                    # If DataFrame has total_price and the total row is zero, skip exporting this sheet
                    if not df.empty and "total_price" in df.columns:
                        try:
                            last_val = pd.to_numeric(df.iloc[-1]["total_price"], errors="coerce")
                            if pd.isna(last_val) or float(last_val) == 0:
                                continue
                        except Exception:
                            pass

                    df.index = range(1, len(df) + 1)
                    df.to_excel(writer, sheet_name=name, index=True, startrow=1)
                    self._style_excel_sheet(writer, name)

                # Append factor sheets ("اقلام ابزار دقیق" and "قیمت نهایی")
                instrument_df, summary_df = self._prepare_factor_sheet(self.tables)

                instrument_df.index = range(1, len(instrument_df) + 1)
                instrument_df.to_excel(writer, sheet_name="اقلام ابزار دقیق", startrow=1, index=False)
                self._style_excel_sheet(writer, "اقلام ابزار دقیق")

                summary_df.index = range(1, len(summary_df) + 1)
                summary_df.to_excel(writer, sheet_name="قیمت نهایی", startrow=1, index=False)
                self._style_excel_sheet(writer, "قیمت نهایی")

            # Open the file that was actually written
            try:
                subprocess.Popen(["start", "", file_path], shell=True)
            except Exception:
                # Best-effort open; non-Windows environments can ignore failures silently
                pass

        except Exception as e:
            show_message(message=str(e), title="Error")

    def _style_excel_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """
        Apply header styling, borders, zebra striping, basic number formatting,
        auto column widths, and freeze the header row.
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        worksheet = writer.sheets[sheet_name]
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # === Styles ===
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
        zebra_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="FFFFAA00", end_color="FFFFAA00", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # === Header row (row 2 because startrow=1) ===
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # === Data rows ===
        for row in worksheet.iter_rows(min_row=3, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = alignment
                cell.border = thin_border

                # Simple price formatting for common numeric columns
                if isinstance(cell.value, (int, float)) and cell.column_letter in ["D", "E", "F", "G"]:
                    cell.number_format = "#,##0"

            # Zebra striping on odd-numbered rows
            if row[0].row % 2 == 1:
                for cell in row:
                    cell.fill = zebra_fill

        # === Highlight last row if it says "جمع کل" (total) ===
        for cell in worksheet[max_row]:
            if isinstance(cell.value, str) and "جمع کل" in cell.value:
                for c in worksheet[max_row]:
                    c.fill = total_fill

        # === Auto column widths ===
        for column_cells in worksheet.columns:
            values = [str(c.value) for c in column_cells if c.value is not None]
            max_length = max((len(v) for v in values), default=0)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 10)

        # === Freeze header row ===
        worksheet.freeze_panes = worksheet["A3"]

    def _prepare_factor_sheet(self, tables: dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Generate the 'instrument items' and 'final price' DataFrames.
        This function preserves the original logic, with additional checks for empty data.
        """
        instruments_list = [
            "Delta Pressure Transmitter",
            "Delta Pressure Switch",
            "Pressure Transmitter",
            "Pressure Switch",
            "Pressure Gauge",
            "Temperature Transmitter",
            "Proximity Switch",
            "Vibration Transmitter",
            "Speed Detector",
            "Level Switch",
            "Level Transmitter",
            "Pt100",
            "Ways Manifold",
            "Calibration",
        ]

        bagfilter_price = 0.0
        instruments_price = 0.0
        cable_price = 0.0
        instrument_items: list[dict] = []

        for name, table in tables.items():
            if name in ("summary_table", "installation_table"):
                continue

            model = table.model()
            if model is None:
                continue

            df_raw = getattr(model, "_data", None)
            if df_raw is None:
                continue

            if isinstance(df_raw, list):
                df = pd.DataFrame(df_raw)
            elif isinstance(df_raw, pd.DataFrame):
                df = df_raw.copy()
            else:
                # Retain original expectation but avoid hard crash
                continue

            # Normalize for lookup; ignore 'brands' column if present
            df = df.drop(columns=["brands"], errors="ignore")
            df.columns = df.columns.str.lower()

            # Handle cable total
            if name == "cable_table":
                if not df.empty and "total_price" in df.columns:
                    try:
                        cable_price = float(pd.to_numeric(df.iloc[-1].get("total_price", 0), errors="coerce") or 0)
                    except Exception:
                        cable_price = 0.0
                continue

            # Aggregate instrument items and panel totals (skip 'Total' rows)
            for _, row in df.iterrows():
                row_type = str(row.get("type", "")).lower()
                if row_type == "total":
                    continue

                total_price_val = float(pd.to_numeric(row.get("total_price", 0), errors="coerce") or 0)
                if any(inst.lower() in row_type for inst in (i.lower() for i in instruments_list)):
                    instruments_price += total_price_val
                    instrument_items.append(
                        {
                            "شرح": row.get("type", ""),
                            "برند": row.get("brand", ""),
                            "تعداد": row.get("quantity", 1),
                            "قیمت واحد(ریال)": row.get("price", 0),
                            "قیمت کل(ریال)": total_price_val,
                        }
                    )

            # Add this panel's total (if present)
            if not df.empty and "total_price" in df.columns:
                try:
                    panel_total = float(pd.to_numeric(df.iloc[-1].get("total_price", 0), errors="coerce") or 0)
                    bagfilter_price += panel_total
                except Exception:
                    pass

        # Remove instrument subtotal from bagfilter aggregate, as in the original logic
        bagfilter_price -= instruments_price

        # Electric motor info (from summary table and specs)
        el_motor_price = 0.0
        el_motor_desc = ""

        summary_model = tables.get("summary_table").model() if tables.get("summary_table") else None
        if summary_model:
            summary_data = getattr(summary_model, "_data", None)
            if isinstance(summary_data, list):
                for row in summary_data:
                    if isinstance(row, dict) and row.get("Title") == "Electric Motor":
                        el_motor_price = float(row.get("Price", 0) or 0)
                        el_motor_desc = row.get("Note", "") or ""
                        break

        fan_motor = self.electrical_specs.get("fan", {}).get("motors", {}).get("fan", {})
        if fan_motor:
            # Build a localized description (preserve original style)
            try:
                kw = int((fan_motor.get("power", 0) or 0) / 1000)
            except Exception:
                kw = 0
            el_motor_desc = (
                f" الکتروموتور {kw}KW "
                f"{fan_motor.get('brand', '')} {fan_motor.get('efficiency_class', '')}"
            )
        elif not el_motor_desc:
            el_motor_desc = "الکتروموتور"

        # Build final summary DataFrame
        summary_data = [
            [0, 0, 0, bagfilter_price, bagfilter_price, 1, "تابلوبگ فیلتر", 1],
            [0, 0, 0, instruments_price, instruments_price, 1, "ابزار دقیق", 2],
            [0, 0, 0, cable_price, cable_price, 1, "کابل", 3],
            [0, 0, 0, 0, 0, 1, "راه اندازی FAT", 4],
            [0, 0, 0, el_motor_price, el_motor_price, 1, el_motor_desc, 5],
        ]

        summary_df = pd.DataFrame(
            summary_data,
            columns=["خرید", "مهندسی", "ساخت", "قیمت کل(ریال)", "قیمت(ریال)", "تعداد", "شرح", "ردیف"],
        )
        # Append totals row
        summary_totals = summary_df[["خرید", "مهندسی", "ساخت", "قیمت کل(ریال)", "قیمت(ریال)"]].sum().to_dict()
        summary_df.loc[len(summary_df)] = {
            **summary_totals,
            "تعداد": "",
            "شرح": "جمع کل(ریال)",
            "ردیف": "",
        }

        instrument_df = pd.DataFrame(instrument_items)

        if not instrument_df.empty:
            instrument_df.insert(0, "ردیف", range(1, len(instrument_df) + 1))
            total_price = float(pd.to_numeric(instrument_df["قیمت کل(ریال)"], errors="coerce").fillna(0).sum())
            instrument_df.loc[len(instrument_df)] = {
                "ردیف": "",
                "شرح": "جمع کل(ریال)",
                "برند": "",
                "تعداد": "",
                "قیمت واحد(ریال)": "",
                "قیمت کل(ریال)": total_price,
            }
            # Reverse columns to match your display preference
            instrument_df = instrument_df[instrument_df.columns[::-1]]

        return instrument_df, summary_df


# -------------------------- Models & Delegates --------------------------

class SummaryTableModel(QAbstractTableModel):
    """
    Editable model for the summary table.
    Columns: Title | Price | Note
    - 'Note' is editable (brand selection); editing updates Price using a brands dict per row.
    """

    def __init__(self, data):
        super().__init__()
        # list of dicts with keys: Title, Price, Note, brands
        self._data = data

    # ---- Qt model overrides ----

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            headers = ["Title", "Price", "Note"]
            if section < len(headers):
                return headers[section]
        return super().headerData(section, orientation, role)

    def rowCount(self, parent: Optional[QModelIndex] = None) -> int:
        return len(self._data)

    def columnCount(self, parent: Optional[QModelIndex] = None) -> int:
        return 3

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):
        if not index.isValid():
            return None

        row, col = index.row(), index.column()
        item = self._data[row]

        if role in (Qt.DisplayRole, Qt.EditRole):
            if col == 0:
                # Title column
                return item.get("Title", "")
            elif col == 1:
                # Price column
                if role == Qt.DisplayRole:
                    # show with thousand separators
                    try:
                        return "{:,}".format(int(item.get("Price", 0)))
                    except Exception:
                        return "{:,}".format(float(item.get("Price", 0) or 0))
                return item.get("Price", 0)
            elif col == 2:
                # Note/Brand column
                return item.get("Note", "")
        return None

    def setData(self, index: QModelIndex, value, role: int = Qt.EditRole) -> bool:
        if not index.isValid() or role != Qt.EditRole:
            return False

        row, col = index.row(), index.column()

        # Allow editing Note (brand) for non-Total rows only
        if col == 2 and row < self.rowCount() - 1:
            item = self._data[row]
            item["Note"] = value
            # Update price from brands dict if available, otherwise keep original
            brands_map = item.get("brands", {})
            if isinstance(brands_map, dict) and value in brands_map:
                item["Price"] = brands_map[value]
            # Emit data changed for Price and Note, and recompute total
            self.dataChanged.emit(self.index(row, 1), self.index(row, 1))
            self.dataChanged.emit(index, index)
            self._update_total()
            return True
        return False

    def flags(self, index: QModelIndex) -> Qt.ItemFlags:
        if not index.isValid():
            return Qt.ItemIsEnabled
        if index.column() == 2 and index.row() < self.rowCount() - 1:
            return Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable

    # ---- helpers ----

    def _update_total(self) -> None:
        """Recalculate the grand total as the sum of all non-Total rows."""
        total_row = self.rowCount() - 1
        try:
            total = sum(float(row.get("Price", 0) or 0) for row in self._data[:-1])
        except Exception:
            total = 0.0
        self._data[total_row]["Price"] = total
        self.dataChanged.emit(self.index(total_row, 1), self.index(total_row, 1))


class BrandComboDelegate(QStyledItemDelegate):
    """Delegate providing a combo box of available brands in the 'Note' column."""

    def createEditor(self, parent, option, index):
        row = index.row()
        brands_dict = index.model()._data[row].get("brands", {})
        if not isinstance(brands_dict, dict) or not brands_dict:
            return None  # no brand options -> no editor
        combo = QComboBox(parent)
        combo.addItems(list(brands_dict.keys()))
        return combo

    def setEditorData(self, editor: QComboBox, index):
        current = index.model().data(index, Qt.EditRole)
        idx = editor.findText(current)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor: QComboBox, model: QAbstractTableModel, index):
        model.setData(index, editor.currentText(), Qt.EditRole)


# -------------------------- Details Tree Viewer --------------------------

class DictionaryTreeViewer(QMainWindow):
    """Simple hierarchical viewer for nested dictionaries with print-to-PDF support."""

    def __init__(self, data: dict, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("Project Details Viewer")
        self.resize(800, 600)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Tree widget
        self.tree = QTreeWidget()
        self.tree.setColumnCount(2)
        self.tree.setHeaderLabels(["Panels", "Value"])
        self.tree.setColumnWidth(0, 300)
        self.populate_tree(self.tree.invisibleRootItem(), data)

        # Buttons
        self.print_button = QPushButton("Print to PDF")
        self.print_button.clicked.connect(self.print_to_pdf)

        self.toggle_button = QPushButton("Expand All")
        self.toggle_button.clicked.connect(self.toggle_tree)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.tree)
        layout.addWidget(self.toggle_button)
        layout.addWidget(self.print_button)
        central_widget.setLayout(layout)

        # State tracker
        self.tree_expanded = False

        self.show()

    def toggle_tree(self) -> None:
        """Toggle between expanded and collapsed tree state."""
        if self.tree_expanded:
            self.tree.collapseAll()
            self.toggle_button.setText("Expand All")
        else:
            self.tree.expandAll()
            self.toggle_button.setText("Collapse All")
        self.tree_expanded = not self.tree_expanded

    def populate_tree(self, parent_item: QTreeWidgetItem, dictionary: dict) -> None:
        """Fill the tree with key/value pairs from a (possibly nested) dict."""

        def format_key(key) -> str:
            return " ".join(word.title() for word in str(key).split("_"))

        for key, value in dictionary.items():
            display_key = format_key(key)
            if isinstance(value, dict):
                item = QTreeWidgetItem([display_key])
                parent_item.addChild(item)
                self.populate_tree(item, value)
            else:
                item = QTreeWidgetItem([display_key, str(value)])
                parent_item.addChild(item)

    def print_to_pdf(self) -> None:
        """Render the entire tree to a single PDF page (scaled to fit)."""
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getSaveFileName(self, "Save PDF", "", "PDF Files (*.pdf)", options=options)
        if not filename:
            return

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(filename)

        self.tree.expandAll()
        self.tree.resizeColumnToContents(0)
        self.tree.resizeColumnToContents(1)

        # Save original size
        original_size = self.tree.size()

        # Estimate total size to capture all rows
        total_height = self.tree.header().height()
        for i in range(self.tree.topLevelItemCount()):
            total_height += self.tree.sizeHintForRow(0) * (1 + self.count_all_items(self.tree.topLevelItem(i)))

        total_width = sum(self.tree.sizeHintForColumn(i) for i in range(self.tree.columnCount())) + 20

        # Temporarily resize tree to fit all contents
        self.tree.resize(total_width, total_height)

        painter = QPainter(printer)

        # Calculate scaling for a single page with margins
        page_rect = printer.pageRect()
        widget_rect = self.tree.rect()

        margin = 50
        scale_x = (page_rect.width() - 2 * margin) / widget_rect.width()
        scale_y = (page_rect.height() - 2 * margin) / widget_rect.height()
        scale = min(scale_x, scale_y)

        painter.translate(page_rect.left() + margin, page_rect.top() + margin)
        painter.scale(scale, scale)

        # Render the tree
        self.tree.render(painter)

        # Optional: thin grid lines for readability
        pen = painter.pen()
        pen.setWidth(1)
        painter.setPen(pen)

        row_height = self.tree.sizeHintForRow(0)
        header_height = self.tree.header().height()
        num_rows = self.count_all_items(self.tree.invisibleRootItem())

        # Horizontal grid lines
        for row in range(num_rows + 2):
            y = header_height + row * row_height
            painter.drawLine(0, y, widget_rect.width(), y)

        # Vertical grid lines
        x = 0
        for col in range(self.tree.columnCount()):
            col_width = self.tree.columnWidth(col)
            x += col_width
            painter.drawLine(x, 0, x, widget_rect.height())

        painter.end()

        # Restore original size
        self.tree.resize(original_size)

    def count_all_items(self, parent_item: QTreeWidgetItem) -> int:
        """Count all descendant items of the given tree item."""
        count = 0
        for i in range(parent_item.childCount()):
            count += 1
            count += self.count_all_items(parent_item.child(i))
        return count
